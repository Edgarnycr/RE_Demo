"""
CRE Underwriting Automation — Demo Sandbox
Streamlit application for OM extraction + Market intelligence + Excel generation.
"""

import io
import json
import re
import traceback
from datetime import datetime
from typing import Optional, Dict, Any

import anthropic
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CRE Underwriting Automation · Demo Sandbox",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# Custom CSS  —  dark navy / gold palette
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    /* ── Global ── */
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main { background: #74a0cf; }

    /* ── Sidebar ── */
    section[data-testid="stSidebar"] {
        background: #0a1628;
        border-right: 1px solid #1e3a5f;
    }
    section[data-testid="stSidebar"] * { color: #c9d6e3 !important; }
    section[data-testid="stSidebar"] .stTextInput input {
        background: #152940;
        border: 1px solid #1e3a5f;
        color: #e8f0fe !important;
        border-radius: 6px;
    }

    /* ── Cards / metrics ── */
    .metric-card {
        background: #101f33;
        border: 1px solid #1e3a5f;
        border-radius: 10px;
        padding: 16px 20px;
        text-align: center;
    }
    .metric-card .label { font-size: 11px; color: #7a9cbf; text-transform: uppercase; letter-spacing: 1px; }
    .metric-card .value { font-size: 18px; font-weight: 700; color: #f0c040; margin-top: 4px; }
    .metric-card .sub   { font-size: 11px; color: #556b7d; margin-top: 2px; }
    .flag-badge {
        display:inline-block;
        background:#3b1f00;
        color:#ffb84d;
        border:1px solid #c47a00;
        border-radius:12px;
        padding:2px 10px;
        font-size:8px;
        margin-left:6px;
    }

    /* ── Headers (dark main area default) ── */
    .main h1, .main h2, .main h3, .main h4 { color: #e8f0fe !important; }
    .hero-title {
        font-size: 2.2rem;
        font-weight: 800;
        color: #f0c040 !important;
        letter-spacing: -0.5px;
    }
    .hero-sub { font-size: 1rem; color: #7a9cbf; margin-top: -8px; }

    /* ── Light theme: darker headings for readability ── */
    [data-testid="stAppViewContainer"][data-theme="light"] .main h1,
    [data-testid="stAppViewContainer"][data-theme="light"] .main h2,
    [data-testid="stAppViewContainer"][data-theme="light"] .main h3,
    [data-testid="stAppViewContainer"][data-theme="light"] .main h4 {
        color: #0a1628 !important;
    }
    [data-testid="stAppViewContainer"][data-theme="light"] .hero-title {
        color: #7a5a00 !important;
    }
    [data-testid="stAppViewContainer"][data-theme="light"] .hero-sub {
        color: #3d4f66 !important;
    }
    html[data-theme="light"] .stApp .main h1,
    html[data-theme="light"] .stApp .main h2,
    html[data-theme="light"] .stApp .main h3,
    html[data-theme="light"] .stApp .main h4 {
        color: #0a1628 !important;
    }
    html[data-theme="light"] .stApp .hero-title { color: #7a5a00 !important; }
    html[data-theme="light"] .stApp .hero-sub { color: #3d4f66 !important; }

    /* ── Expander ── */
    .streamlit-expanderHeader {
        background: #101f33 !important;
        color: #c9d6e3 !important;
        border: 1px solid #1e3a5f !important;
        border-radius: 8px !important;
    }

    /* ── Buttons ── */
    .stButton > button {
        background: linear-gradient(135deg, #c9a227, #f0c040);
        color: #0a1628;
        font-weight: 700;
        border: none;
        border-radius: 8px;
        padding: 10px 28px;
        font-size: 15px;
        transition: opacity .2s;
    }
    .stButton > button:hover { opacity: 0.88; }

    /* ── Upload zone ── */
    .stFileUploader > div { border: 2px dashed #1e3a5f !important; border-radius: 10px !important; background: #0d1b2a !important; }

    /* ── Divider ── */
    hr { border-color: #1e3a5f !important; }

    /* ── Sidebar "tab" selector (rectangular pills) ── */
    section[data-testid="stSidebar"] div[role="radiogroup"] {
        gap: 0.45rem;
        margin: 4px 0 4px 0;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] > label[data-baseweb="radio"] {
        background: #152940;
        border: 1px solid #1e3a5f;
        border-radius: 10px;
        padding: 10px 10px;
        margin: 0px 4px 8px 4px;
        transition: all .15s ease-in-out;
    }
    section[data-testid="stSidebar"] div[role="radiogroup"] > label[data-baseweb="radio"]:hover {
        border-color: #f0c040;
        background: #12253b;
    }
    /* hide the little radio circle */
    section[data-testid="stSidebar"] label[data-baseweb="radio"] > div:first-child {
        display: none !important;
    }
    /* selected state */
    section[data-testid="stSidebar"] label[data-baseweb="radio"]:has(div[role="radio"][aria-checked="true"]) {
        border-color: #f0c040;
        background: linear-gradient(135deg, rgba(201,162,39,0.22), rgba(240,192,64,0.12));
        box-shadow: 0 0 0 1px rgba(240,192,64,0.25) inset;
    }
    /* fallback selected state (when :has isn't supported) */
    section[data-testid="stSidebar"] div[role="radio"][aria-checked="true"] {
        background: transparent !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ─────────────────────────────────────────────────────────────────────────────
# SECRETS — API keys (server) + invite-only access
# ─────────────────────────────────────────────────────────────────────────────
def _secrets_get(key: str, default: str = "") -> str:
    try:
        val = st.secrets[key]
        if val is None:
            return default
        return str(val).strip()
    except Exception:
        return default


def _get_invite_codes() -> set[str]:
    """
    Valid invite passwords from Streamlit secrets.
    Use APP_INVITE_CODES as a TOML list, or a single comma-separated string,
    or a single APP_INVITE_CODE string.
    """
    codes: set[str] = set()
    try:
        if "APP_INVITE_CODES" in st.secrets:
            raw = st.secrets["APP_INVITE_CODES"]
            if isinstance(raw, (list, tuple)):
                codes = {str(x).strip() for x in raw if str(x).strip()}
            else:
                s = str(raw).strip()
                if s:
                    codes = {p.strip() for p in s.split(",") if p.strip()}
        if "APP_INVITE_CODE" in st.secrets:
            one = str(st.secrets["APP_INVITE_CODE"]).strip()
            if one:
                codes.add(one)
    except Exception:
        pass
    return codes


def _is_valid_invite(code: str) -> bool:
    if not code or not code.strip():
        return False
    invite_codes = _get_invite_codes()
    if not invite_codes:
        return False
    return code.strip() in invite_codes


def _admin_access_code() -> str:
    return _secrets_get("ADMIN_ACCESS_CODE", "EDGAR_ADMIN")


def effective_anthropic_key() -> str:
    return _secrets_get("ANTHROPIC_API_KEY") or (st.session_state.get("anthropic_key") or "")


def effective_perplexity_key() -> str:
    return _secrets_get("PERPLEXITY_API_KEY") or (st.session_state.get("perplexity_key") or "")


INVITE_ONLY_MSG = (
    "This app is **invite-only**. Enter the access code you were given in **Client ID / Access Code** "
    "in the sidebar to use this section."
)


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR  —  Client auth + Navigation + Developer tools
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    # Invite-only access (password in this field unlocks the app for the session)
    st.markdown("### Client Access")
    client_id = st.text_input(
        "Client ID / Access Code",
        type="password",
        help="Enter the invite code you were given. This is required to use the app.",
    )
    entered = (client_id or "").strip()
    if entered:
        if _is_valid_invite(entered):
            st.session_state["access_granted"] = True
            st.session_state["user_folder"] = entered
        else:
            st.session_state["access_granted"] = False
            st.session_state["user_folder"] = None
            st.sidebar.error("Invalid access code.")
    else:
        st.session_state["access_granted"] = False
        st.session_state["user_folder"] = None

    if not _get_invite_codes():
        st.caption("Deploy: set `APP_INVITE_CODES` (and API keys) in Streamlit Secrets.")

    st.markdown("---")

    st.markdown("## Navigation")
    page = st.radio(
        "Page",
        ["Underwriting", "Portfolio Management", "Market Intelligence", "Developer Tools"],
        label_visibility="collapsed",
        key="page",
    )
    st.markdown("---")

    # Ensure keys exist in session even when not editing them
    if "anthropic_key" not in st.session_state:
        st.session_state["anthropic_key"] = ""
    if "perplexity_key" not in st.session_state:
        st.session_state["perplexity_key"] = ""

# ─────────────────────────────────────────────────────────────────────────────
# HERO HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<p class="hero-title">🏦 CRE Underwriting Automation</p>', unsafe_allow_html=True)
st.markdown('<p class="hero-sub">Demo Sandbox · Institutional Underwriting & Market Intelligence</p>', unsafe_allow_html=True)
st.markdown("---")

if page == "Underwriting":
    st.markdown(
        """
        <div style='font-size:12px;color:#556b7d;line-height:1.7'>
        <b style='color:#7a9cbf'>How it works</b><br>
        1. Upload a Broker OM (PDF)<br>
        2. Claude extracts financial data<br>
        3. Perplexity or Claude adds market intel<br>
        4. Review both datasets<br>
        5. Export an Institutional Excel Memo
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION PROMPT
# ─────────────────────────────────────────────────────────────────────────────
EXTRACTION_PROMPT = """You are an expert real estate financial analyst. Extract structured data from the following Broker Offering Memorandum text.

IMPORTANT INSTRUCTIONS:
- Only extract values explicitly stated or clearly inferable.
- Do NOT hallucinate or guess missing values.
- If a field is not found, return null.
- Normalize all numeric values (remove commas, $, %, etc.).
- Convert percentages to decimals (e.g., 65% → 0.65).
- Ensure all numbers are floats or integers (not strings).
- Be conservative — prioritize accuracy over completeness.
- Also extract the city and state of the property.

CONFIDENCE RULES:
For each extracted field, assess confidence (0–100%) internally.
If confidence < 25%: add companion field with suffix _confidence_flag = "Low confidence"
If confidence ≥ 25%: do NOT include a confidence flag.

FIELDS TO EXTRACT:
- property_name (string)
- property_type (string) — Multifamily, Industrial, Office, Retail, Hotel
- city (string)
- state (string) — full state name or 2-letter abbreviation
- units (integer)
- total_sf (number)
- deal_type (string) — Loan Origination | Refinancing | Construction Loan | Mezzanine Loan | Equity Acquisition
- noi (number) — annual NOI
- source_noi_type (string) — Stabilized | T12 | Underwritten
- purchase_price (number)
- loan_amount (number)
- ltv (number) — decimal
- dscr (number)
- loan_scenario_selected (string)
- rent_roll_summary (object, optional):
    - total_units (integer)
    - occupied_units (integer)
    - occupancy_rate (number, decimal)
    - average_rent (number)

NOI PRIORITY: Stabilized > Underwritten > T12. If ambiguous, use most conservative.
LTV: use stated value, else calculate loan_amount / purchase_price.
DEAL TYPE: Refinance/take-out → Refinancing; Acquisition/purchase → Loan Origination; Ground-up/development → Construction Loan; Mezz/preferred equity → Mezzanine Loan; Equity raise/JV → Equity Acquisition.
Multiple loan scenarios: prefer Base Case, else highest loan amount.
Total SF: prefer NRA or GLA.

OUTPUT: Return ONLY valid JSON. No markdown. No code blocks. Start with { end with }.

Base structure:
{
  "property_name": "",
  "property_type": "",
  "city": "",
  "state": "",
  "units": null,
  "total_sf": null,
  "deal_type": "",
  "noi": null,
  "source_noi_type": "",
  "purchase_price": null,
  "loan_amount": null,
  "ltv": null,
  "dscr": null,
  "loan_scenario_selected": "",
  "rent_roll_summary": {
    "total_units": null,
    "occupied_units": null,
    "occupancy_rate": null,
    "average_rent": null
  }
}

OM TEXT:
"""

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def extract_pdf_text(pdf_bytes: bytes) -> str:
    """Extract raw text from PDF bytes using pypdf."""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for page in reader.pages:
        txt = page.extract_text()
        if txt:
            pages.append(txt)
    return "\n\n".join(pages)


def extract_with_anthropic(pdf_text: str, api_key: str) -> dict:
    """Call Claude 3.5 Sonnet to extract structured financial data."""
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=2048,
        messages=[
            {
                "role": "user",
                "content": EXTRACTION_PROMPT + pdf_text[:80000],  # token guard
            }
        ],
    )
    raw = message.content[0].text.strip()
    # strip any accidental markdown fences
    raw = re.sub(r"^```(?:json)?", "", raw).strip()
    raw = re.sub(r"```$", "", raw).strip()
    return json.loads(raw)

def compute_extraction_scores(extracted: dict) -> tuple[int, int, str]:
    """
    Returns: (completion_pct, confidence_score, suggested_prompt_addition)
    """
    key_fields = [
        "property_name",
        "property_type",
        "city",
        "state",
        "units",
        "total_sf",
        "deal_type",
        "noi",
        "source_noi_type",
        "purchase_price",
        "loan_amount",
        "ltv",
        "dscr",
        "loan_scenario_selected",
    ]

    def _is_present(v):
        if v is None:
            return False
        if isinstance(v, str) and not v.strip():
            return False
        return True

    present = sum(1 for f in key_fields if _is_present(extracted.get(f)))
    completion_pct = int(round(100 * present / max(len(key_fields), 1)))

    low_flags = [
        k
        for k, v in extracted.items()
        if k.endswith("_confidence_flag") and v == "Low confidence"
    ]
    missing_core = [
        f
        for f in ["city", "state", "property_type", "noi", "purchase_price", "loan_amount"]
        if not _is_present(extracted.get(f))
    ]

    confidence_score = 100
    confidence_score -= min(40, 8 * len(low_flags))
    confidence_score -= min(30, 6 * len(missing_core))
    confidence_score = max(0, min(100, confidence_score))

    suggestions = []
    if missing_core:
        suggestions.append(
            "Ask the OM to explicitly state: "
            + ", ".join(missing_core)
            + " (place near executive summary)."
        )
    if low_flags:
        flagged_fields = sorted({k.replace("_confidence_flag", "") for k in low_flags})
        suggestions.append(
            "Add instruction: if values are in tables, read exact label+value for "
            + ", ".join(flagged_fields[:6])
            + ("…" if len(flagged_fields) > 6 else "")
            + "."
        )
    if not suggestions:
        suggestions.append("No prompt change suggested — extraction looked consistent.")

    return completion_pct, confidence_score, " ".join(suggestions)


def get_perplexity_research(city: str, state: str, property_type: str, api_key: str) -> dict:
    """Call Perplexity Sonar to get structured market intelligence."""
    query = (
        f"Provide a structured market report for {property_type} in {city}, {state} for Q1 2026. "
        f"I need: "
        f"1. Market Occupancy Rate, "
        f"2. Average PSF Asking Rent, "
        f"3. Net Absorption (last 12 months), "
        f"4. Total SF Under Construction. "
        f"Also, list 5-10 recent Sales Comps for {property_type} in this market including "
        f"Property Name/Address, Sale Price, Sale Date, and SF if available."
    )

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": "sonar",
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are a commercial real estate market analyst. "
                    "Always structure your response with clear sections: "
                    "MARKET METRICS and SALES COMPS. "
                    "Be specific with numbers and cite sources where possible."
                ),
            },
            {"role": "user", "content": query},
        ],
        "max_tokens": 2000,
    }

    resp = requests.post(
        "https://api.perplexity.ai/chat/completions",
        headers=headers,
        json=payload,
        timeout=45,
    )
    resp.raise_for_status()
    data = resp.json()
    raw_text = data["choices"][0]["message"]["content"]
    citations = data.get("citations", [])

    return parse_perplexity_response(raw_text, citations, city, state, property_type)


def get_claude_market_summary(
    city: str, state: str, property_type: str, extracted: dict, api_key: str
) -> dict:
    """Brief market + submarket trends with subject-property context via Claude (no live web data)."""
    client = anthropic.Anthropic(api_key=api_key)

    rr = extracted.get("rent_roll_summary") or {}
    subject_ctx = {
        "property_name": extracted.get("property_name"),
        "property_type": extracted.get("property_type"),
        "location": f"{extracted.get('city','')}, {extracted.get('state','')}",
        "units": extracted.get("units"),
        "total_sf": extracted.get("total_sf"),
        "noi": extracted.get("noi"),
        "purchase_price": extracted.get("purchase_price"),
        "loan_amount": extracted.get("loan_amount"),
        "ltv": extracted.get("ltv"),
        "dscr": extracted.get("dscr"),
        "occupancy_rate": rr.get("occupancy_rate"),
        "avg_rent": rr.get("average_rent"),
    }

    prompt = (
        f"You are a commercial real estate market analyst. For {property_type} in {city}, {state}, write a brief "
        "professional summary (under 450 words) with exactly two sections:\n\n"
        "## MARKET TRENDS\n"
        "Regional / metro (CBSA) trends for this asset type: demand, rent direction, supply pipeline, capital markets. "
        "Use general industry knowledge; do not invent specific statistics; do not cite sources.\n"
        "Include 2-3 bullets of what these trends imply for underwriting (risk/neutral/upside).\n\n"
        "## SUBMARKET TRENDS\n"
        "Local submarket dynamics in/near the subject location: competition set, tenant demand, new deliveries, and "
        "near-term risks/opportunities.\n"
        "Weave in SUBJECT PROPERTY context using only the extracted OM fields below. If a field is missing, say "
        "\"Not disclosed\" and avoid guessing.\n"
        "Explicitly address:\n"
        "- How the subject appears to be performing today (occupancy/rents/NOI/debt metrics if provided)\n"
        "- Whether there is potential to add value (operational, leasing, repositioning, capex) and/or adjust pricing "
        "(if performance vs implied underwriting looks tight)\n\n"
        "Extracted OM fields (may contain nulls):\n"
        f"{json.dumps(subject_ctx, ensure_ascii=False)}"
    )
    message = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=2048,
        messages=[{"role": "user", "content": prompt}],
    )
    raw_text = message.content[0].text.strip()
    return {
        "raw_text": raw_text,
        "metrics": {
            "market_occupancy": None,
            "avg_psf_rent": None,
            "net_absorption": None,
            "sf_under_construction": None,
        },
        "comps": [],
        "citations": ["Claude — general knowledge (not live web data)"],
        "city": city,
        "state": state,
        "property_type": property_type,
        "market_source": "claude_fallback",
    }


def parse_perplexity_response(text: str, citations: list, city: str, state: str, property_type: str) -> dict:
    """Parse Perplexity free-form text into structured metrics + comps."""

    metrics = {
        "market_occupancy": None,
        "avg_psf_rent": None,
        "net_absorption": None,
        "sf_under_construction": None,
    }

    occ_match = re.search(
        r"(?:occupancy|vacancy)[^\d]*(\d+(?:\.\d+)?)\s*%", text, re.IGNORECASE
    )
    if occ_match:
        val = float(occ_match.group(1))
        # if it looks like vacancy, invert
        if "vacanc" in text[max(0, occ_match.start() - 20):occ_match.start()].lower():
            val = round(100 - val, 2)
        metrics["market_occupancy"] = f"{val}%"

    psf_match = re.search(
        r"\$\s*(\d+(?:\.\d+)?)\s*(?:per|\/)\s*(?:sf|sqft|square\s*foot)", text, re.IGNORECASE
    )
    if psf_match:
        metrics["avg_psf_rent"] = f"${psf_match.group(1)} PSF"

    abs_match = re.search(
        r"net\s+absorption[^\d\-]*([+-]?\s*[\d,]+(?:\.\d+)?)\s*(?:sf|sq\.?\s*ft)", text, re.IGNORECASE
    )
    if abs_match:
        metrics["net_absorption"] = abs_match.group(1).replace(",", "").strip() + " SF"

    const_match = re.search(
        r"(?:under\s+construction|construction\s+pipeline)[^\d]*([\d,]+(?:\.\d+)?)\s*(?:sf|sq\.?\s*ft|msf|million)",
        text,
        re.IGNORECASE,
    )
    if const_match:
        metrics["sf_under_construction"] = const_match.group(1).replace(",", "").strip() + " SF"

    # ── Sales comps ──────────────────────────────────────────────────────────
    comps = []
    # Try to find a comps section
    comps_section_match = re.search(
        r"(?:sales?\s*comps?|recent\s+(?:sales?|transactions?))[:\s]*(.*)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    if comps_section_match:
        comps_text = comps_section_match.group(1)
        # Split on numbered list items or bullet points
        entries = re.split(r"\n\s*(?:\d+[\.\)]\s+|\*\s+|-\s+)", comps_text)
        for entry in entries[:10]:
            entry = entry.strip()
            if len(entry) < 10:
                continue
            price_m = re.search(r"\$\s*([\d,]+(?:\.\d+)?)\s*(?:million|M\b)?", entry, re.I)
            date_m = re.search(r"((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-,]*\d{4}|\d{1,2}[\/\-]\d{4}|Q[1-4]\s*\d{4})", entry, re.I)
            sf_m = re.search(r"([\d,]+)\s*(?:sf|sq\.?\s*ft)", entry, re.I)

            comp = {
                "description": entry[:120],
                "price": price_m.group(0) if price_m else "N/A",
                "date": date_m.group(0) if date_m else "N/A",
                "sf": sf_m.group(0) if sf_m else "N/A",
            }
            comps.append(comp)

    return {
        "raw_text": text,
        "metrics": metrics,
        "comps": comps,
        "citations": citations,
        "city": city,
        "state": state,
        "property_type": property_type,
        "market_source": "perplexity",
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def _style_header_row(ws, row, cols, fill_hex="1e3a5f", font_hex="F0C040"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_hex, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="2d5986")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def _style_data_row(ws, row, cols, even=False):
    bg = "0f1e30" if not even else "101f33"
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color="c9d6e3", size=10)
    align = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin", color="1a3050")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def fmt_currency(val):
    if val is None:
        return "N/A"
    try:
        v = float(val)
        if v >= 1_000_000:
            return f"${v/1_000_000:.2f}M"
        return f"${v:,.0f}"
    except Exception:
        return str(val)


def fmt_percent(val):
    if val is None:
        return "N/A"
    try:
        v = float(val)
        if v <= 1:
            return f"{v*100:.1f}%"
        return f"{v:.1f}%"
    except Exception:
        return str(val)


def fmt_number(val):
    if val is None:
        return "N/A"
    try:
        return f"{float(val):,.0f}"
    except Exception:
        return str(val)


def build_excel(extracted: dict, market: dict) -> bytes:
    """Build institutional-grade Excel workbook in memory."""
    wb = Workbook()

    # ── TAB 1: Internal Underwriting ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Internal Underwriting"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "1e3a5f"

    # Tab background
    ws1.sheet_view.showGridLines = False
    for col in range(1, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 28

    # Title block
    ws1.merge_cells("A1:D1")
    title_cell = ws1["A1"]
    title_cell.value = "INTERNAL UNDERWRITING SUMMARY"
    title_cell.font = Font(bold=True, color="F0C040", size=14)
    title_cell.fill = PatternFill("solid", fgColor="0a1628")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:D2")
    sub_cell = ws1["A2"]
    prop = extracted.get("property_name", "N/A") or "N/A"
    sub_cell.value = f"Property: {prop}  |  Deal Type: {extracted.get('deal_type','N/A') or 'N/A'}"
    sub_cell.font = Font(color="7a9cbf", size=10, italic=True)
    sub_cell.fill = PatternFill("solid", fgColor="0a1628")
    sub_cell.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 20

    ws1.row_dimensions[3].height = 10  # spacer

    # Column headers
    headers = ["Field", "Value", "Confidence Flag", "Notes"]
    for i, h in enumerate(headers, 1):
        ws1.cell(row=4, column=i).value = h
    _style_header_row(ws1, 4, 4)
    ws1.row_dimensions[4].height = 22

    # Data rows
    flag_icon = "⚠ Low Confidence"

    def _flag(field):
        return flag_icon if extracted.get(f"{field}_confidence_flag") == "Low confidence" else ""

    rows = [
        ("Property Name",       extracted.get("property_name") or "N/A",           _flag("property_name"),     ""),
        ("Property Type",       extracted.get("property_type") or "N/A",           _flag("property_type"),     ""),
        ("Location",            f"{extracted.get('city','')}, {extracted.get('state','')}", "", ""),
        ("Units",               fmt_number(extracted.get("units")),                 _flag("units"),             ""),
        ("Total SF",            fmt_number(extracted.get("total_sf")),              _flag("total_sf"),          "NRA / GLA"),
        ("Deal Type",           extracted.get("deal_type") or "N/A",               _flag("deal_type"),         ""),
        ("NOI",                 fmt_currency(extracted.get("noi")),                 _flag("noi"),               "Annualized"),
        ("NOI Source",          extracted.get("source_noi_type") or "N/A",         _flag("source_noi_type"),   "Stabilized → T12 priority"),
        ("Purchase Price",      fmt_currency(extracted.get("purchase_price")),      _flag("purchase_price"),    ""),
        ("Loan Amount",         fmt_currency(extracted.get("loan_amount")),         _flag("loan_amount"),       ""),
        ("LTV",                 fmt_percent(extracted.get("ltv")),                  _flag("ltv"),               ""),
        ("DSCR",                fmt_number(extracted.get("dscr")) if extracted.get("dscr") else "N/A",
                                                                                    _flag("dscr"),              ""),
        ("Loan Scenario",       extracted.get("loan_scenario_selected") or "N/A",  "", ""),
    ]

    rr = extracted.get("rent_roll_summary") or {}
    if any(rr.get(k) for k in ["total_units", "occupied_units", "occupancy_rate", "average_rent"]):
        rows.append(("— RENT ROLL —", "", "", ""))
        rows.append(("Total Units",    fmt_number(rr.get("total_units")),      _flag("rent_roll_summary.total_units"),    ""))
        rows.append(("Occupied Units", fmt_number(rr.get("occupied_units")),   _flag("rent_roll_summary.occupied_units"), ""))
        rows.append(("Occupancy Rate", fmt_percent(rr.get("occupancy_rate")), _flag("rent_roll_summary.occupancy_rate"), ""))
        rows.append(("Avg Monthly Rent", fmt_currency(rr.get("average_rent")), _flag("rent_roll_summary.average_rent"), "Per Unit"))

    for idx, (field, val, flag, note) in enumerate(rows):
        r = idx + 5
        ws1.cell(row=r, column=1).value = field
        ws1.cell(row=r, column=2).value = val
        ws1.cell(row=r, column=3).value = flag
        ws1.cell(row=r, column=4).value = note

        if field.startswith("—"):
            # Section divider
            for col in range(1, 5):
                c = ws1.cell(row=r, column=col)
                c.fill = PatternFill("solid", fgColor="1e3a5f")
                c.font = Font(bold=True, color="F0C040", size=10)
                c.alignment = Alignment(horizontal="center")
        else:
            _style_data_row(ws1, r, 4, even=(idx % 2 == 0))
            if flag:
                ws1.cell(row=r, column=3).font = Font(color="FFB84D", bold=True, size=10)
        ws1.row_dimensions[r].height = 18

    # ── TAB 2: Market Intel ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Market Intel")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "c9a227"
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 40

    # Title
    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    loc = f"{market.get('city','')}, {market.get('state','')} — {market.get('property_type','')}"
    t2.value = f"MARKET INTELLIGENCE REPORT  |  {loc}  |  Q1 2026"
    t2.font = Font(bold=True, color="F0C040", size=13)
    t2.fill = PatternFill("solid", fgColor="0a1628")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.row_dimensions[2].height = 10

    # Market Metrics header
    for i, h in enumerate(["Market Metric", "Value", "Source / Notes"], 1):
        ws2.cell(row=3, column=i).value = h
    _style_header_row(ws2, 3, 3)
    ws2.row_dimensions[3].height = 22

    metrics_map = {
        "Market Occupancy Rate": market["metrics"].get("market_occupancy", "See raw report"),
        "Average PSF Asking Rent": market["metrics"].get("avg_psf_rent", "See raw report"),
        "Net Absorption (12 mo)": market["metrics"].get("net_absorption", "See raw report"),
        "Total SF Under Construction": market["metrics"].get("sf_under_construction", "See raw report"),
    }
    src = market.get("market_source")
    if src == "claude_fallback":
        cit_str = "Claude — general knowledge (not live web data)"
    else:
        cit_str = ", ".join(market.get("citations", [])[:3]) or (
            "Perplexity Sonar" if src == "perplexity" else "N/A"
        )

    for idx, (metric, val) in enumerate(metrics_map.items()):
        r = idx + 4
        ws2.cell(row=r, column=1).value = metric
        ws2.cell(row=r, column=2).value = val or "N/A"
        ws2.cell(row=r, column=3).value = cit_str
        _style_data_row(ws2, r, 3, even=(idx % 2 == 0))
        ws2.row_dimensions[r].height = 18

    # Spacer
    ws2.row_dimensions[8].height = 12

    # Sales comps sub-header
    ws2.merge_cells("A9:C9")
    sc = ws2["A9"]
    sc.value = "RECENT SALES COMPARABLES"
    sc.font = Font(bold=True, color="F0C040", size=11)
    sc.fill = PatternFill("solid", fgColor="0a1628")
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[9].height = 22

    for i, h in enumerate(["Property / Description", "Sale Price", "Date  |  SF"], 1):
        ws2.cell(row=10, column=i).value = h
    _style_header_row(ws2, 10, 3)

    comps = market.get("comps", [])
    if comps:
        for idx, comp in enumerate(comps[:10]):
            r = idx + 11
            ws2.cell(row=r, column=1).value = comp.get("description", "N/A")
            ws2.cell(row=r, column=2).value = comp.get("price", "N/A")
            ws2.cell(row=r, column=3).value = f"{comp.get('date','N/A')}  |  {comp.get('sf','N/A')}"
            _style_data_row(ws2, r, 3, even=(idx % 2 == 0))
            ws2.row_dimensions[r].height = 18
    else:
        ws2.cell(row=11, column=1).value = "No structured comps parsed — see Raw Market Report tab"
        _style_data_row(ws2, 11, 3)

    # ── TAB 3: Raw Market Report ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Raw Market Report")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "2d5986"
    ws3.column_dimensions["A"].width = 120

    ws3.merge_cells("A1:A1")
    t3 = ws3["A1"]
    _src = market.get("market_source")
    if _src == "claude_fallback":
        t3.value = "CLAUDE MARKET / SUBMARKET SUMMARY"
    elif _src == "perplexity":
        t3.value = "RAW PERPLEXITY SONAR RESPONSE"
    else:
        t3.value = "RAW MARKET REPORT"
    t3.font = Font(bold=True, color="F0C040", size=13)
    t3.fill = PatternFill("solid", fgColor="0a1628")
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 28

    ws3.row_dimensions[2].height = 8

    raw = market.get("raw_text", "No data received.")
    for line_idx, line in enumerate(raw.split("\n")):
        r = line_idx + 3
        c = ws3.cell(row=r, column=1)
        c.value = line
        c.font = Font(color="c9d6e3", size=10)
        c.fill = PatternFill("solid", fgColor="0d1b2a")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[r].height = 15

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for key in [
    "extracted",
    "market",
    "excel_bytes",
    "pdf_text",
    "file_log",
    "stored_files",
    "user_folder",
    "access_granted",
]:
    if key not in st.session_state:
        if key in {"file_log", "stored_files"}:
            st.session_state[key] = []
        elif key == "access_granted":
            st.session_state[key] = False
        else:
            st.session_state[key] = None

# ─────────────────────────────────────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────────────────────────────────────
def _store_uploaded_pdf(uploaded_file_obj) -> Optional[Dict[str, Any]]:
    if uploaded_file_obj is None:
        return None
    if not st.session_state.get("access_granted"):
        return None
    pdf_bytes = uploaded_file_obj.getvalue()
    file_id = f"{uploaded_file_obj.name}:{len(pdf_bytes)}"
    existing_ids = {f.get("id") for f in (st.session_state.get("stored_files") or [])}
    if file_id in existing_ids:
        return next((f for f in st.session_state["stored_files"] if f.get("id") == file_id), None)
    entry = {
        "id": file_id,
        "timestamp": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "filename": uploaded_file_obj.name,
        "pdf_bytes": pdf_bytes,
        "owner": st.session_state.get("user_folder"),
    }
    st.session_state["stored_files"].append(entry)
    st.session_state["stored_files"] = st.session_state["stored_files"][-20:]
    return entry


if page == "Underwriting":
    if not st.session_state.get("access_granted"):
        st.warning(INVITE_ONLY_MSG + " Upload and extraction require a valid invite.", icon="🔒")
    else:
        anthropic_key = effective_anthropic_key()
        perplexity_key = effective_perplexity_key()

        st.markdown("### 📄 Upload Broker Offering Memorandum")
        uploaded_file = st.file_uploader(
            "Drag & drop or click to upload a PDF",
            type=["pdf"],
            label_visibility="collapsed",
            key="broker_om_uploader",
        )
        stored_entry = _store_uploaded_pdf(uploaded_file)

        if uploaded_file and anthropic_key:
            col_run, col_info = st.columns([1, 3])
            with col_run:
                run_btn = st.button("⚙️ Extract & Analyze", use_container_width=True)
            with col_info:
                if not perplexity_key:
                    st.info(
                        "Perplexity key not set — a brief market & submarket summary will be generated with Claude "
                        "(general knowledge, not live web data).",
                        icon="ℹ️",
                    )

            if run_btn:
                with st.status("🔍 Processing your OM…", expanded=True) as status:

                    # 1. PDF text
                    st.write("📃 Extracting PDF text…")
                    pdf_bytes = stored_entry["pdf_bytes"] if stored_entry else uploaded_file.getvalue()
                    pdf_text = extract_pdf_text(pdf_bytes)
                    st.session_state["pdf_text"] = pdf_text

                    # 2. Anthropic extraction
                    st.write("🤖 Sending to Claude 3.5 Sonnet for structured extraction…")
                    try:
                        extracted = extract_with_anthropic(pdf_text, anthropic_key)
                        st.session_state["extracted"] = extracted
                        st.write(
                            f"✅ Extracted {len([k for k, v in extracted.items() if v is not None])} fields."
                        )
                    except Exception as e:
                        st.error(f"Anthropic extraction failed: {e}")
                        st.stop()

                    # 3. Market intelligence (Perplexity, or Claude fallback)
                    city = extracted.get("city", "") or ""
                    state = extracted.get("state", "") or ""
                    ptype = extracted.get("property_type", "Commercial") or "Commercial"

                    if city and state:
                        if perplexity_key:
                            st.write(
                                f"🌐 Fetching market intelligence for {ptype} in {city}, {state}…"
                            )
                            try:
                                market = get_perplexity_research(city, state, ptype, perplexity_key)
                                st.session_state["market"] = market
                                st.write(
                                    f"✅ Market data received. {len(market.get('comps', []))} comps parsed."
                                )
                            except Exception as e:
                                st.warning(
                                    f"Perplexity unavailable ({e}). Using Claude for market & submarket summary."
                                )
                                try:
                                    market = get_claude_market_summary(
                                        city, state, ptype, extracted, anthropic_key
                                    )
                                    st.session_state["market"] = market
                                    st.write("✅ Claude market summary ready (general knowledge, not live data).")
                                except Exception as ce:
                                    st.error(f"Claude market summary failed: {ce}")
                                    st.session_state["market"] = {
                                        "raw_text": f"Perplexity error: {e}\nClaude fallback error: {ce}",
                                        "metrics": {},
                                        "comps": [],
                                        "citations": [],
                                        "city": city,
                                        "state": state,
                                        "property_type": ptype,
                                        "market_source": "error",
                                    }
                        else:
                            st.write(
                                f"🤖 Generating market & submarket summary with Claude for {ptype} in {city}, {state}…"
                            )
                            try:
                                market = get_claude_market_summary(
                                    city, state, ptype, extracted, anthropic_key
                                )
                                st.session_state["market"] = market
                                st.write("✅ Claude market summary ready (general knowledge, not live data).")
                            except Exception as e:
                                st.error(f"Claude market summary failed: {e}")
                                st.session_state["market"] = None
                    else:
                        st.warning("City/State not found in OM — skipping market research.")
                        st.session_state["market"] = None

                    st.session_state["excel_bytes"] = None  # reset on new run
                    status.update(label="✅ Analysis complete!", state="complete")
                    completion_pct, confidence_score, suggestion = compute_extraction_scores(extracted)
                    owner = st.session_state.get("user_folder")
                    st.session_state["file_log"].append(
                        {
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "Filename": uploaded_file.name,
                            "Property Name (Extracted)": extracted.get("property_name") or "N/A",
                            "Confidence Score": f"{confidence_score}% (completion {completion_pct}%)",
                            "Suggested Prompt Addition/correction": suggestion,
                            "Owner": owner,
                        }
                    )
                    st.session_state["file_log"] = st.session_state["file_log"][-50:]

        elif uploaded_file and not anthropic_key:
            st.warning(
                "Anthropic API key is not configured. Add **ANTHROPIC_API_KEY** to Streamlit Secrets (or use "
                "**Developer Tools** locally with a valid invite code).",
                icon="🔑",
            )

        st.markdown("---")
        st.markdown("### 🗂️ File Log")
        st.markdown(
            "Uploaded PDFs are stored in-memory for this session and can be downloaded any time."
        )

        user_folder = st.session_state.get("user_folder")
        is_admin = user_folder == _admin_access_code()

        if not user_folder:
            st.info("Enter a valid invite access code in the sidebar to view your files.")
        else:
            # Filter stored files by owner unless admin
            if st.session_state["stored_files"]:
                if is_admin:
                    visible_files = list(st.session_state["stored_files"])
                else:
                    visible_files = [
                        f for f in st.session_state["stored_files"] if f.get("owner") == user_folder
                    ]

                if visible_files:
                    for f in reversed(visible_files):
                        c1, c2, c3 = st.columns([3, 2, 1])
                        c1.markdown(f"**{f.get('filename','(untitled)')}**")
                        c2.markdown(
                            f"<span style='color:#7a9cbf'>{f.get('timestamp','')}</span>",
                            unsafe_allow_html=True,
                        )
                        c3.download_button(
                            "Download",
                            data=f.get("pdf_bytes", b""),
                            file_name=f.get("filename", "uploaded.pdf"),
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"dl_{f.get('id')}",
                        )
                else:
                    st.info("No PDFs uploaded yet for this invite code.")

            else:
                st.info("No PDFs uploaded yet.")

            st.markdown("#### Extraction Activity")
            if st.session_state["file_log"]:
                log_df = pd.DataFrame(st.session_state["file_log"])
                if not is_admin:
                    log_df = (
                        log_df[log_df["Owner"] == user_folder]
                        if "Owner" in log_df.columns
                        else log_df.iloc[0:0]
                    )

                if not log_df.empty:
                    log_df = log_df[
                        [
                            "Timestamp",
                            "Filename",
                            "Property Name (Extracted)",
                            "Confidence Score",
                            "Suggested Prompt Addition/correction",
                        ]
                    ]
                    st.dataframe(log_df, use_container_width=True, hide_index=True)
                else:
                    st.info("No extractions logged yet for this invite code.")
            else:
                st.info("No extractions logged yet.")

elif page == "Portfolio Management":
    if not st.session_state.get("access_granted"):
        st.warning(INVITE_ONLY_MSG, icon="🔒")
    else:
        st.markdown("### 📈 Portfolio Management (Placeholder Dashboard)")

        pm_data = pd.DataFrame(
            [
                {"Client": "Client A", "Property Type": "Multi Family", "Region": "Northeast", "AUM": 450_000_000, "Commitments": 120_000_000, "LTV": 0.63, "DSCR": 1.48, "DY": 0.092},
                {"Client": "Client A", "Property Type": "Industrial", "Region": "South", "AUM": 310_000_000, "Commitments": 80_000_000, "LTV": 0.58, "DSCR": 1.62, "DY": 0.087},
                {"Client": "Client B", "Property Type": "Office", "Region": "West", "AUM": 220_000_000, "Commitments": 60_000_000, "LTV": 0.67, "DSCR": 1.28, "DY": 0.101},
                {"Client": "Client B", "Property Type": "Hospitality", "Region": "Midwest", "AUM": 140_000_000, "Commitments": 35_000_000, "LTV": 0.61, "DSCR": 1.35, "DY": 0.112},
                {"Client": "Client C", "Property Type": "Multi Family", "Region": "South", "AUM": 510_000_000, "Commitments": 150_000_000, "LTV": 0.60, "DSCR": 1.55, "DY": 0.089},
                {"Client": "Client C", "Property Type": "Industrial", "Region": "West", "AUM": 275_000_000, "Commitments": 70_000_000, "LTV": 0.56, "DSCR": 1.70, "DY": 0.082},
            ]
        )

        base_total_aum = float(pm_data["AUM"].sum())
        scaling_factor = 3_200_000_000.0 / base_total_aum if base_total_aum else 1.0

        f1, f2, f3, f4 = st.columns(4)
        with f1:
            clients = ["All", "Client A", "Client B", "Client C"]
            sel_client = st.selectbox("Clients", clients, index=0)
        with f2:
            ptypes = ["All", "Multi Family", "Industrial", "Office", "Hospitality"]
            sel_ptype = st.selectbox("Property type", ptypes, index=0)
        with f3:
            regions = ["All", "Northeast", "South", "Midwest", "West"]
            sel_region = st.selectbox("Region", regions, index=0)
        with f4:
            status_opts = ["All", "Active", "Watch"]
            sel_status = st.selectbox("Status", status_opts, index=0)

        df = pm_data.copy()
        if sel_client != "All":
            df = df[df["Client"] == sel_client]
        if sel_ptype != "All":
            df = df[df["Property Type"] == sel_ptype]
        if sel_region != "All":
            df = df[df["Region"] == sel_region]
        # status is a placeholder filter in this demo; keep it interactive without changing data
        _ = sel_status

        import altair as alt

        def _donut_chart(counts: pd.DataFrame, title: str, color_col: str):
            base = (
                alt.Chart(counts)
                .mark_arc(innerRadius=55, outerRadius=90)
                .encode(
                    theta=alt.Theta("count:Q"),
                    color=alt.Color(
                        f"{color_col}:N",
                        scale=alt.Scale(
                            range=["#f0c040", "#2d5986", "#7a9cbf", "#c9a227", "#1e3a5f"]
                        ),
                        legend=None,
                    ),
                    tooltip=[color_col, "count"],
                )
                .properties(title=title, width=260, height=230)
                .configure_title(anchor="middle", fontSize=13, dy=10)
            )
            return base

        if df.empty:
            st.warning("No results for the selected filters.")
            prop_counts = pd.DataFrame({"Property Type": ["N/A"], "count": [1]})
            region_counts = pd.DataFrame({"Region": ["N/A"], "count": [1]})
        else:
            prop_counts = df.groupby("Property Type", as_index=False).size().rename(columns={"size": "count"})
            region_counts = df.groupby("Region", as_index=False).size().rename(columns={"size": "count"})

        st.markdown("<br>", unsafe_allow_html=True)
        c_chart1, c_chart2, c_summary = st.columns([1, 1, 1.1], vertical_alignment="top")
        with c_chart1:
            st.markdown(
                "<div style='text-align:center'><h4 style='margin-bottom:4px'>Property type</h4></div>",
                unsafe_allow_html=True,
            )
            st.altair_chart(_donut_chart(prop_counts, "", "Property Type"), use_container_width=True)
        with c_chart2:
            st.markdown(
                "<div style='text-align:center'><h4 style='margin-bottom:4px'>Region</h4></div>",
                unsafe_allow_html=True,
            )
            st.altair_chart(_donut_chart(region_counts, "", "Region"), use_container_width=True)
        with c_summary:
            raw_total_aum = float(df["AUM"].sum()) if not df.empty else 0.0
            total_aum = raw_total_aum * scaling_factor if raw_total_aum else 0.0
            avg_ltv = float((df["LTV"] * df["AUM"]).sum() / max(df["AUM"].sum(), 1)) if not df.empty else 0.0
            w_dscr = float((df["DSCR"] * df["AUM"]).sum() / max(df["AUM"].sum(), 1)) if not df.empty else 0.0
            w_dy = float((df["DY"] * df["AUM"]).sum() / max(df["AUM"].sum(), 1)) if not df.empty else 0.0

            summary = pd.DataFrame(
                {
                    "Metric": [
                        "Total AUM",
                        "Total Loans",
                        "Average LTV",
                        "Weighted DSCR",
                        "Weighted DY",
                    ],
                    "Value": [
                        fmt_currency(total_aum),
                        "80",
                        fmt_percent(avg_ltv),
                        f"{w_dscr:.2f}" if w_dscr else "N/A",
                        fmt_percent(w_dy),
                    ],
                }
            )
            st.markdown("#### Summary")
            st.dataframe(summary, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("#### Watch Monitor")
        watch_df = pd.DataFrame(
            [
                {
                    "Loan name": "Sunset Apartments – Senior Loan",
                    "Balance": "$38,500,000",
                    "Property type": "Multi Family",
                    "LTV": "66%",
                    "DSCR": "1.25x",
                    "DY": "10.1%",
                    "Subject Property Occupancy": "92%",
                    "Market Occupancy": "94%",
                    "FWD Yr rollover": "18%",
                },
                {
                    "Loan name": "Riverbend Logistics – Bridge",
                    "Balance": "$52,000,000",
                    "Property type": "Industrial",
                    "LTV": "58%",
                    "DSCR": "1.48x",
                    "DY": "8.6%",
                    "Subject Property Occupancy": "96%",
                    "Market Occupancy": "95%",
                    "FWD Yr rollover": "9%",
                },
                {
                    "Loan name": "Central Plaza – Refi",
                    "Balance": "$44,250,000",
                    "Property type": "Office",
                    "LTV": "71%",
                    "DSCR": "1.12x",
                    "DY": "11.4%",
                    "Subject Property Occupancy": "84%",
                    "Market Occupancy": "88%",
                    "FWD Yr rollover": "27%",
                },
            ]
        )
        if sel_ptype != "All":
            watch_df = watch_df[watch_df["Property type"] == sel_ptype]
        st.dataframe(
            watch_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Loan name": st.column_config.TextColumn(width=280),
            },
        )

        st.markdown("#### All Commitments")
        all_commit_cols = list(watch_df.columns)
        all_commit_df = pd.DataFrame(columns=all_commit_cols)
        st.dataframe(
            all_commit_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Loan name": st.column_config.TextColumn(width=280),
            },
        )

elif page == "Market Intelligence":
    if not st.session_state.get("access_granted"):
        st.warning(INVITE_ONLY_MSG, icon="🔒")
    else:
        st.markdown("### 🌐 Market Intelligence")

        hdr_zip, hdr_ptype, hdr_btn = st.columns([1.1, 1.2, 0.7], vertical_alignment="bottom")
        with hdr_zip:
            st.markdown("**Postal Code**")
        with hdr_ptype:
            st.markdown("**Property type**")
        with hdr_btn:
            st.markdown("")  # align with label row above Run button

        row_zip, row_ptype, row_btn = st.columns([1.1, 1.2, 0.7], vertical_alignment="center")
        with row_zip:
            mi_postal = st.text_input(
                "postal_code_mi",
                value="",
                placeholder="Enter Here",
                label_visibility="collapsed",
            )
        with row_ptype:
            mi_ptype = st.selectbox(
                "property_type_mi",
                ["Multi Family", "Industrial", "Office", "Hospitality"],
                index=0,
                label_visibility="collapsed",
            )
        with row_btn:
            mi_run = st.button("Run", use_container_width=True)

        if mi_run and mi_postal.strip():
            st.markdown("---")
            st.markdown("#### Market & Submarket Trends")
            st.markdown(
                f"""
**Location**: {mi_postal.strip()}  ·  **Property type**: {mi_ptype}

- **Current market conditions**: Leasing velocity is healthy with stable to modestly rising asking rents, while capital remains selective and focused on well-sponsored transactions. Lender underwriting is emphasizing in-place cash flow, rollover timing, and business-plan execution risk.
- **Submarket dynamics**: Nearby assets are competing primarily on amenity quality and tenant improvement packages, with new supply clustered around the most transit- and amenity-rich nodes. Occupancy at institutional assets is trending slightly above the broader market, but backfill risk exists for older, less renovated product.
- **Forward-looking view (12–24 months)**: Base case assumes muted but positive rent growth, modest cap-rate expansion, and continued bifurcation between prime locations and secondary corridors. Downside scenarios focus on tighter debt markets and slower lease-up; upside scenarios depend on faster-than-expected absorption and limited incremental supply.
"""
            )

            st.markdown("#### Relevant Assets")
            relevant_df = pd.DataFrame(
                [
                    {
                        "Loan name": "Subject – Benchmark 1",
                        "Balance": "$42,000,000",
                        "Property type": mi_ptype,
                        "LTV": "63%",
                        "DSCR": "1.35x",
                        "DY": "9.4%",
                        "Subject Property Occupancy": "95%",
                        "Market Occupancy": "93%",
                        "FWD Yr rollover": "14%",
                    },
                    {
                        "Loan name": "Subject – Benchmark 2",
                        "Balance": "$35,500,000",
                        "Property type": mi_ptype,
                        "LTV": "59%",
                        "DSCR": "1.42x",
                        "DY": "8.9%",
                        "Subject Property Occupancy": "92%",
                        "Market Occupancy": "91%",
                        "FWD Yr rollover": "11%",
                    },
                ]
            )
            st.dataframe(
                relevant_df,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Loan name": st.column_config.TextColumn(width=280),
                },
            )
elif page == "Developer Tools":
    if not st.session_state.get("access_granted"):
        st.warning(INVITE_ONLY_MSG, icon="🔒")
    else:
        st.markdown("### 🔧 Developer Tools")
        st.markdown(
            "Optional overrides when **ANTHROPIC_API_KEY** / **PERPLEXITY_API_KEY** are not set in "
            "Streamlit Secrets. In production, prefer Secrets only."
        )
        st.markdown("---")

        anthropic_key = st.text_input(
            "Anthropic API Key (override)",
            type="password",
            placeholder="sk-ant-...",
            value=st.session_state.get("anthropic_key", ""),
            help="Used only if ANTHROPIC_API_KEY is not set in Streamlit Secrets",
        )
        perplexity_key = st.text_input(
            "Perplexity API Key (override)",
            type="password",
            placeholder="pplx-...",
            value=st.session_state.get("perplexity_key", ""),
            help="Used only if PERPLEXITY_API_KEY is not set in Streamlit Secrets",
        )
        st.session_state["anthropic_key"] = anthropic_key
        st.session_state["perplexity_key"] = perplexity_key

        st.markdown("---")
        st.markdown(
            "<div style='font-size:12px;color:#556b7d;line-height:1.6'>"
            "Invite codes and API keys for production should live in <b>Streamlit Secrets</b>."
            "</div>",
            unsafe_allow_html=True,
        )

# ─────────────────────────────────────────────────────────────────────────────
# RESULTS
# ─────────────────────────────────────────────────────────────────────────────
if page == "Underwriting" and st.session_state.get("access_granted") and st.session_state["extracted"]:
    extracted = st.session_state["extracted"]
    market    = st.session_state["market"]

    st.markdown("---")
    st.markdown("### 📊 Extraction Results")

    # ── Quick KPI row ────────────────────────────────────────────────────────
    kpis = [
        ("Property Type",    extracted.get("property_type") or "N/A",          ""),
        ("Deal Type",        extracted.get("deal_type") or "N/A",               ""),
        ("Purchase Price",   fmt_currency(extracted.get("purchase_price")),     ""),
        ("Loan Amount",      fmt_currency(extracted.get("loan_amount")),        ""),
        ("LTV",              fmt_percent(extracted.get("ltv")),                 ""),
        ("NOI",              fmt_currency(extracted.get("noi")),                extracted.get("source_noi_type") or ""),
    ]
    cols = st.columns(len(kpis))
    for col, (label, value, sub) in zip(cols, kpis):
        flag = extracted.get(f"{label.lower().replace(' ','_')}_confidence_flag","")
        badge = '<span class="flag-badge">⚠ Low</span>' if flag else ""
        col.markdown(
            f'<div class="metric-card">'
            f'<div class="label">{label}{badge}</div>'
            f'<div class="value">{value}</div>'
            f'<div class="sub">{sub}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Expanders ────────────────────────────────────────────────────────────
    with st.expander("🔍 Full Extracted PDF Data (review before export)", expanded=False):
        # highlight confidence flags
        clean = {k: v for k, v in extracted.items() if not k.endswith("_confidence_flag")}
        flags = {k: v for k, v in extracted.items() if k.endswith("_confidence_flag")}

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Extracted Fields**")
            st.json(clean)
        with c2:
            if flags:
                st.markdown("**⚠ Confidence Flags**")
                st.json(flags)
            else:
                st.success("No low-confidence fields detected.", icon="✅")

    if market:
        _mkt_title = (
            "🤖 Claude Market & Submarket Summary (review before export)"
            if market.get("market_source") == "claude_fallback"
            else "🌐 Perplexity Market Intelligence (review before export)"
        )
        with st.expander(_mkt_title, expanded=False):
            m1, m2 = st.columns(2)
            with m1:
                st.markdown("**Key Market Metrics**")
                if market.get("metrics"):
                    for k, v in market["metrics"].items():
                        label = k.replace("_", " ").title()
                        st.markdown(f"- **{label}:** {v or 'N/A'}")
                else:
                    st.info("Metrics could not be parsed from response.")

                if market.get("comps"):
                    st.markdown(f"**{len(market['comps'])} Sales Comps Found**")
                    comp_df = pd.DataFrame(market["comps"])
                    st.dataframe(comp_df, use_container_width=True, hide_index=True)

            with m2:
                st.markdown(
                    "**Full text**"
                    if market.get("market_source") == "claude_fallback"
                    else "**Raw Sonar Response**"
                )
                st.markdown(
                    f'<div style="background:#0d1b2a;border:1px solid #1e3a5f;border-radius:8px;'
                    f'padding:14px;max-height:340px;overflow-y:auto;font-size:12px;'
                    f'color:#c9d6e3;white-space:pre-wrap;">{market.get("raw_text","")}</div>',
                    unsafe_allow_html=True,
                )

    # ── Generate Excel ────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Export Institutional Memo")

    gen_col, dl_col = st.columns([1, 2])
    with gen_col:
        if st.button("📊 Generate Institutional Memo", use_container_width=True):
            if not market:
                # create empty market stub
                market = {
                    "raw_text": "Market data not available.",
                    "metrics": {},
                    "comps": [],
                    "citations": [],
                    "city": extracted.get("city", ""),
                    "state": extracted.get("state", ""),
                    "property_type": extracted.get("property_type", ""),
                    "market_source": "none",
                }
            with st.spinner("Building Excel workbook…"):
                excel_bytes = build_excel(extracted, market)
                st.session_state["excel_bytes"] = excel_bytes
            st.success("Excel workbook ready for download!", icon="✅")

    with dl_col:
        if st.session_state["excel_bytes"]:
            prop_slug = re.sub(r"[^a-zA-Z0-9]", "_", extracted.get("property_name","OM") or "OM")[:30]
            st.download_button(
                label="⬇️  Download Institutional Memo (.xlsx)",
                data=st.session_state["excel_bytes"],
                file_name=f"{prop_slug}_Sizer.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ─────────────────────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<div style="text-align:center;color:#3d5570;font-size:12px;">'
    "CRE Automation · Demo Sandbox · Powered by Claude &amp; Perplexity · Zero-Retention Architecture"
    "</div>",
    unsafe_allow_html=True,
)
