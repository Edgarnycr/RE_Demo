"""
Real Estate Debt Fund — Demo Sandbox
Streamlit application for OM extraction + market intelligence + Excel generation.
"""

import io
import json
import re
import traceback

import anthropic
import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

# ─────────────────────────────────────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CRE Underwriting Automation · Demo Sandbox",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# Custom CSS  —  dark navy / gold palette
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    /* ── Global ── */
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .main { background: #0d1b2a; }

    /* ── Sidebar ── */
    section[data-testid="stSidebar"] {
        background: #0a1628;
        border-right: 1px solid #1e3a5f;
    }
    section[data-testid="stSidebar"] * { color: #c9d6e3 !important; }
    section[data-testid="stSidebar"] .stTextInput input {
        background: #152940;
        border: 1px solid #1e3a5f;
        color: #e8f0fe !important;
        border-radius: 6px;
    }

    /* ── Cards / metrics ── */
    .metric-card {
        background: #101f33;
        border: 1px solid #1e3a5f;
        border-radius: 10px;
        padding: 16px 20px;
        text-align: center;
    }
    .metric-card .label { font-size: 11px; color: #7a9cbf; text-transform: uppercase; letter-spacing: 1px; }
    .metric-card .value { font-size: 24px; font-weight: 700; color: #f0c040; margin-top: 4px; }
    .metric-card .sub   { font-size: 11px; color: #556b7d; margin-top: 2px; }
    .flag-badge {
        display:inline-block;
        background:#3b1f00;
        color:#ffb84d;
        border:1px solid #c47a00;
        border-radius:12px;
        padding:2px 10px;
        font-size:11px;
        margin-left:6px;
    }

    /* ── Headers ── */
    h1, h2, h3 { color: #e8f0fe !important; }
    .hero-title {
        font-size: 2.2rem;
        font-weight: 800;
        color: #f0c040 !important;
        letter-spacing: -0.5px;
    }
    .hero-sub { font-size: 1rem; color: #7a9cbf; margin-top: -8px; }

    /* ── Expander ── */
    .streamlit-expanderHeader {
        background: #101f33 !important;
        color: #c9d6e3 !important;
        border: 1px solid #1e3a5f !important;
        border-radius: 8px !important;
    }

    /* ── Buttons ── */
    .stButton > button {
        background: linear-gradient(135deg, #c9a227, #f0c040);
        color: #0a1628;
        font-weight: 700;
        border: none;
        border-radius: 8px;
        padding: 10px 28px;
        font-size: 15px;
        transition: opacity .2s;
    }
    .stButton > button:hover { opacity: 0.88; }

    /* ── Upload zone ── */
    .stFileUploader > div { border: 2px dashed #1e3a5f !important; border-radius: 10px !important; background: #0d1b2a !important; }

    /* ── Divider ── */
    hr { border-color: #1e3a5f !important; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR  —  API Keys & info
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔑 API Configuration")
    st.markdown("---")
    anthropic_key = st.text_input(
        "Anthropic API Key",
        type="password",
        placeholder="sk-ant-...",
        help="Used for PDF extraction via Claude 3.5 Sonnet",
    )
    perplexity_key = st.text_input(
        "Perplexity API Key",
        type="password",
        placeholder="pplx-...",
        help="Used for live market intelligence via Sonar",
    )
    st.markdown("---")
    st.markdown(
        """
        <div style='font-size:12px;color:#556b7d;line-height:1.7'>
        <b style='color:#7a9cbf'>How it works</b><br>
        1. Upload a Broker OM (PDF)<br>
        2. Claude extracts financial data<br>
        3. Perplexity enriches with market intel<br>
        4. Review both datasets<br>
        5. Export an Institutional Excel Memo
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("---")
    st.markdown(
        "<div style='font-size:11px;color:#3d5570'>Zero-retention: no files are written to disk.</div>",
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────────────────────────────────────
# HERO HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<p class="hero-title">🏦 Real Estate Debt Fund</p>', unsafe_allow_html=True)
st.markdown('<p class="hero-sub">Demo Sandbox · Institutional Underwriting & Market Intelligence</p>', unsafe_allow_html=True)
st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION PROMPT
# ─────────────────────────────────────────────────────────────────────────────
EXTRACTION_PROMPT = """You are an expert real estate financial analyst. Extract structured data from the following Broker Offering Memorandum text.

IMPORTANT INSTRUCTIONS:
- Only extract values explicitly stated or clearly inferable.
- Do NOT hallucinate or guess missing values.
- If a field is not found, return null.
- Normalize all numeric values (remove commas, $, %, etc.).
- Convert percentages to decimals (e.g., 65% → 0.65).
- Ensure all numbers are floats or integers (not strings).
- Be conservative — prioritize accuracy over completeness.
- Also extract the city and state of the property.

CONFIDENCE RULES:
For each extracted field, assess confidence (0–100%) internally.
If confidence < 25%: add companion field with suffix _confidence_flag = "Low confidence"
If confidence ≥ 25%: do NOT include a confidence flag.

FIELDS TO EXTRACT:
- property_name (string)
- property_type (string) — Multifamily, Industrial, Office, Retail, Hotel
- city (string)
- state (string) — full state name or 2-letter abbreviation
- units (integer)
- total_sf (number)
- deal_type (string) — Loan Origination | Refinancing | Construction Loan | Mezzanine Loan | Equity Acquisition
- noi (number) — annual NOI
- source_noi_type (string) — Stabilized | T12 | Underwritten
- purchase_price (number)
- loan_amount (number)
- ltv (number) — decimal
- dscr (number)
- loan_scenario_selected (string)
- rent_roll_summary (object, optional):
    - total_units (integer)
    - occupied_units (integer)
    - occupancy_rate (number, decimal)
    - average_rent (number)

NOI PRIORITY: Stabilized > Underwritten > T12. If ambiguous, use most conservative.
LTV: use stated value, else calculate loan_amount / purchase_price.
DEAL TYPE: Refinance/take-out → Refinancing; Acquisition/purchase → Loan Origination; Ground-up/development → Construction Loan; Mezz/preferred equity → Mezzanine Loan; Equity raise/JV → Equity Acquisition.
Multiple loan scenarios: prefer Base Case, else highest loan amount.
Total SF: prefer NRA or GLA.

OUTPUT: Return ONLY valid JSON. No markdown. No code blocks. Start with { end with }.

Base structure:
{
  "property_name": "",
  "property_type": "",
  "city": "",
  "state": "",
  "units": null,
  "total_sf": null,
  "deal_type": "",
  "noi": null,
  "source_noi_type": "",
  "purchase_price": null,
  "loan_amount": null,
  "ltv": null,
  "dscr": null,
  "loan_scenario_selected": "",
  "rent_roll_summary": {
    "total_units": null,
    "occupied_units": null,
    "occupancy_rate": null,
    "average_rent": null
  }
}

OM TEXT:
"""

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def extract_pdf_text(pdf_bytes: bytes) -> str:
    """Extract raw text from PDF bytes using pypdf."""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    pages = []
    for page in reader.pages:
        txt = page.extract_text()
        if txt:
            pages.append(txt)
    return "\n\n".join(pages)


def extract_with_anthropic(pdf_text: str, api_key: str) -> dict:
    """Call Claude 3.5 Sonnet to extract structured financial data."""
    client = anthropic.Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=2048,
        messages=[
            {
                "role": "user",
                "content": EXTRACTION_PROMPT + pdf_text[:80000],  # token guard
            }
        ],
    )
    raw = message.content[0].text.strip()
    # strip any accidental markdown fences
    raw = re.sub(r"^```(?:json)?", "", raw).strip()
    raw = re.sub(r"```$", "", raw).strip()
    return json.loads(raw)


def get_perplexity_research(city: str, state: str, property_type: str, api_key: str) -> dict:
    """Call Perplexity Sonar to get structured market intelligence."""
    query = (
        f"Provide a structured market report for {property_type} in {city}, {state} for Q1 2026. "
        f"I need: "
        f"1. Market Occupancy Rate, "
        f"2. Average PSF Asking Rent, "
        f"3. Net Absorption (last 12 months), "
        f"4. Total SF Under Construction. "
        f"Also, list 5-10 recent Sales Comps for {property_type} in this market including "
        f"Property Name/Address, Sale Price, Sale Date, and SF if available."
    )

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": "sonar",
        "messages": [
            {
                "role": "system",
                "content": (
                    "You are a commercial real estate market analyst. "
                    "Always structure your response with clear sections: "
                    "MARKET METRICS and SALES COMPS. "
                    "Be specific with numbers and cite sources where possible."
                ),
            },
            {"role": "user", "content": query},
        ],
        "max_tokens": 2000,
    }

    resp = requests.post(
        "https://api.perplexity.ai/chat/completions",
        headers=headers,
        json=payload,
        timeout=45,
    )
    resp.raise_for_status()
    data = resp.json()
    raw_text = data["choices"][0]["message"]["content"]
    citations = data.get("citations", [])

    return parse_perplexity_response(raw_text, citations, city, state, property_type)


def parse_perplexity_response(text: str, citations: list, city: str, state: str, property_type: str) -> dict:
    """Parse Perplexity free-form text into structured metrics + comps."""

    metrics = {
        "market_occupancy": None,
        "avg_psf_rent": None,
        "net_absorption": None,
        "sf_under_construction": None,
    }

    occ_match = re.search(
        r"(?:occupancy|vacancy)[^\d]*(\d+(?:\.\d+)?)\s*%", text, re.IGNORECASE
    )
    if occ_match:
        val = float(occ_match.group(1))
        # if it looks like vacancy, invert
        if "vacanc" in text[max(0, occ_match.start() - 20):occ_match.start()].lower():
            val = round(100 - val, 2)
        metrics["market_occupancy"] = f"{val}%"

    psf_match = re.search(
        r"\$\s*(\d+(?:\.\d+)?)\s*(?:per|\/)\s*(?:sf|sqft|square\s*foot)", text, re.IGNORECASE
    )
    if psf_match:
        metrics["avg_psf_rent"] = f"${psf_match.group(1)} PSF"

    abs_match = re.search(
        r"net\s+absorption[^\d\-]*([+-]?\s*[\d,]+(?:\.\d+)?)\s*(?:sf|sq\.?\s*ft)", text, re.IGNORECASE
    )
    if abs_match:
        metrics["net_absorption"] = abs_match.group(1).replace(",", "").strip() + " SF"

    const_match = re.search(
        r"(?:under\s+construction|construction\s+pipeline)[^\d]*([\d,]+(?:\.\d+)?)\s*(?:sf|sq\.?\s*ft|msf|million)",
        text,
        re.IGNORECASE,
    )
    if const_match:
        metrics["sf_under_construction"] = const_match.group(1).replace(",", "").strip() + " SF"

    # ── Sales comps ──────────────────────────────────────────────────────────
    comps = []
    # Try to find a comps section
    comps_section_match = re.search(
        r"(?:sales?\s*comps?|recent\s+(?:sales?|transactions?))[:\s]*(.*)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    if comps_section_match:
        comps_text = comps_section_match.group(1)
        # Split on numbered list items or bullet points
        entries = re.split(r"\n\s*(?:\d+[\.\)]\s+|\*\s+|-\s+)", comps_text)
        for entry in entries[:10]:
            entry = entry.strip()
            if len(entry) < 10:
                continue
            price_m = re.search(r"\$\s*([\d,]+(?:\.\d+)?)\s*(?:million|M\b)?", entry, re.I)
            date_m = re.search(r"((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-,]*\d{4}|\d{1,2}[\/\-]\d{4}|Q[1-4]\s*\d{4})", entry, re.I)
            sf_m = re.search(r"([\d,]+)\s*(?:sf|sq\.?\s*ft)", entry, re.I)

            comp = {
                "description": entry[:120],
                "price": price_m.group(0) if price_m else "N/A",
                "date": date_m.group(0) if date_m else "N/A",
                "sf": sf_m.group(0) if sf_m else "N/A",
            }
            comps.append(comp)

    return {
        "raw_text": text,
        "metrics": metrics,
        "comps": comps,
        "citations": citations,
        "city": city,
        "state": state,
        "property_type": property_type,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def _style_header_row(ws, row, cols, fill_hex="1e3a5f", font_hex="F0C040"):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color=font_hex, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="2d5986")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def _style_data_row(ws, row, cols, even=False):
    bg = "0f1e30" if not even else "101f33"
    fill = PatternFill("solid", fgColor=bg)
    font = Font(color="c9d6e3", size=10)
    align = Alignment(vertical="center", wrap_text=True)
    thin = Side(style="thin", color="1a3050")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def fmt_currency(val):
    if val is None:
        return "N/A"
    try:
        v = float(val)
        if v >= 1_000_000:
            return f"${v/1_000_000:.2f}M"
        return f"${v:,.0f}"
    except Exception:
        return str(val)


def fmt_percent(val):
    if val is None:
        return "N/A"
    try:
        v = float(val)
        if v <= 1:
            return f"{v*100:.1f}%"
        return f"{v:.1f}%"
    except Exception:
        return str(val)


def fmt_number(val):
    if val is None:
        return "N/A"
    try:
        return f"{float(val):,.0f}"
    except Exception:
        return str(val)


def build_excel(extracted: dict, market: dict) -> bytes:
    """Build institutional-grade Excel workbook in memory."""
    wb = Workbook()

    # ── TAB 1: Internal Underwriting ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Internal Underwriting"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "1e3a5f"

    # Tab background
    ws1.sheet_view.showGridLines = False
    for col in range(1, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 28

    # Title block
    ws1.merge_cells("A1:D1")
    title_cell = ws1["A1"]
    title_cell.value = "INTERNAL UNDERWRITING SUMMARY"
    title_cell.font = Font(bold=True, color="F0C040", size=14)
    title_cell.fill = PatternFill("solid", fgColor="0a1628")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:D2")
    sub_cell = ws1["A2"]
    prop = extracted.get("property_name", "N/A") or "N/A"
    sub_cell.value = f"Property: {prop}  |  Deal Type: {extracted.get('deal_type','N/A') or 'N/A'}"
    sub_cell.font = Font(color="7a9cbf", size=10, italic=True)
    sub_cell.fill = PatternFill("solid", fgColor="0a1628")
    sub_cell.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 20

    ws1.row_dimensions[3].height = 10  # spacer

    # Column headers
    headers = ["Field", "Value", "Confidence Flag", "Notes"]
    for i, h in enumerate(headers, 1):
        ws1.cell(row=4, column=i).value = h
    _style_header_row(ws1, 4, 4)
    ws1.row_dimensions[4].height = 22

    # Data rows
    flag_icon = "⚠ Low Confidence"

    def _flag(field):
        return flag_icon if extracted.get(f"{field}_confidence_flag") == "Low confidence" else ""

    rows = [
        ("Property Name",       extracted.get("property_name") or "N/A",           _flag("property_name"),     ""),
        ("Property Type",       extracted.get("property_type") or "N/A",           _flag("property_type"),     ""),
        ("Location",            f"{extracted.get('city','')}, {extracted.get('state','')}", "", ""),
        ("Units",               fmt_number(extracted.get("units")),                 _flag("units"),             ""),
        ("Total SF",            fmt_number(extracted.get("total_sf")),              _flag("total_sf"),          "NRA / GLA"),
        ("Deal Type",           extracted.get("deal_type") or "N/A",               _flag("deal_type"),         ""),
        ("NOI",                 fmt_currency(extracted.get("noi")),                 _flag("noi"),               "Annualized"),
        ("NOI Source",          extracted.get("source_noi_type") or "N/A",         _flag("source_noi_type"),   "Stabilized → T12 priority"),
        ("Purchase Price",      fmt_currency(extracted.get("purchase_price")),      _flag("purchase_price"),    ""),
        ("Loan Amount",         fmt_currency(extracted.get("loan_amount")),         _flag("loan_amount"),       ""),
        ("LTV",                 fmt_percent(extracted.get("ltv")),                  _flag("ltv"),               ""),
        ("DSCR",                fmt_number(extracted.get("dscr")) if extracted.get("dscr") else "N/A",
                                                                                    _flag("dscr"),              ""),
        ("Loan Scenario",       extracted.get("loan_scenario_selected") or "N/A",  "", ""),
    ]

    rr = extracted.get("rent_roll_summary") or {}
    if any(rr.get(k) for k in ["total_units", "occupied_units", "occupancy_rate", "average_rent"]):
        rows.append(("— RENT ROLL —", "", "", ""))
        rows.append(("Total Units",    fmt_number(rr.get("total_units")),      _flag("rent_roll_summary.total_units"),    ""))
        rows.append(("Occupied Units", fmt_number(rr.get("occupied_units")),   _flag("rent_roll_summary.occupied_units"), ""))
        rows.append(("Occupancy Rate", fmt_percent(rr.get("occupancy_rate")), _flag("rent_roll_summary.occupancy_rate"), ""))
        rows.append(("Avg Monthly Rent", fmt_currency(rr.get("average_rent")), _flag("rent_roll_summary.average_rent"), "Per Unit"))

    for idx, (field, val, flag, note) in enumerate(rows):
        r = idx + 5
        ws1.cell(row=r, column=1).value = field
        ws1.cell(row=r, column=2).value = val
        ws1.cell(row=r, column=3).value = flag
        ws1.cell(row=r, column=4).value = note

        if field.startswith("—"):
            # Section divider
            for col in range(1, 5):
                c = ws1.cell(row=r, column=col)
                c.fill = PatternFill("solid", fgColor="1e3a5f")
                c.font = Font(bold=True, color="F0C040", size=10)
                c.alignment = Alignment(horizontal="center")
        else:
            _style_data_row(ws1, r, 4, even=(idx % 2 == 0))
            if flag:
                ws1.cell(row=r, column=3).font = Font(color="FFB84D", bold=True, size=10)
        ws1.row_dimensions[r].height = 18

    # ── TAB 2: Market Intel ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Market Intel")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "c9a227"
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 40

    # Title
    ws2.merge_cells("A1:C1")
    t2 = ws2["A1"]
    loc = f"{market.get('city','')}, {market.get('state','')} — {market.get('property_type','')}"
    t2.value = f"MARKET INTELLIGENCE REPORT  |  {loc}  |  Q1 2026"
    t2.font = Font(bold=True, color="F0C040", size=13)
    t2.fill = PatternFill("solid", fgColor="0a1628")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.row_dimensions[2].height = 10

    # Market Metrics header
    for i, h in enumerate(["Market Metric", "Value", "Source / Notes"], 1):
        ws2.cell(row=3, column=i).value = h
    _style_header_row(ws2, 3, 3)
    ws2.row_dimensions[3].height = 22

    metrics_map = {
        "Market Occupancy Rate": market["metrics"].get("market_occupancy", "See raw report"),
        "Average PSF Asking Rent": market["metrics"].get("avg_psf_rent", "See raw report"),
        "Net Absorption (12 mo)": market["metrics"].get("net_absorption", "See raw report"),
        "Total SF Under Construction": market["metrics"].get("sf_under_construction", "See raw report"),
    }
    cit_str = ", ".join(market.get("citations", [])[:3]) or "Perplexity Sonar"

    for idx, (metric, val) in enumerate(metrics_map.items()):
        r = idx + 4
        ws2.cell(row=r, column=1).value = metric
        ws2.cell(row=r, column=2).value = val or "N/A"
        ws2.cell(row=r, column=3).value = cit_str
        _style_data_row(ws2, r, 3, even=(idx % 2 == 0))
        ws2.row_dimensions[r].height = 18

    # Spacer
    ws2.row_dimensions[8].height = 12

    # Sales comps sub-header
    ws2.merge_cells("A9:C9")
    sc = ws2["A9"]
    sc.value = "RECENT SALES COMPARABLES"
    sc.font = Font(bold=True, color="F0C040", size=11)
    sc.fill = PatternFill("solid", fgColor="0a1628")
    sc.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[9].height = 22

    for i, h in enumerate(["Property / Description", "Sale Price", "Date  |  SF"], 1):
        ws2.cell(row=10, column=i).value = h
    _style_header_row(ws2, 10, 3)

    comps = market.get("comps", [])
    if comps:
        for idx, comp in enumerate(comps[:10]):
            r = idx + 11
            ws2.cell(row=r, column=1).value = comp.get("description", "N/A")
            ws2.cell(row=r, column=2).value = comp.get("price", "N/A")
            ws2.cell(row=r, column=3).value = f"{comp.get('date','N/A')}  |  {comp.get('sf','N/A')}"
            _style_data_row(ws2, r, 3, even=(idx % 2 == 0))
            ws2.row_dimensions[r].height = 18
    else:
        ws2.cell(row=11, column=1).value = "No structured comps parsed — see Raw Market Report tab"
        _style_data_row(ws2, 11, 3)

    # ── TAB 3: Raw Market Report ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Raw Market Report")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "2d5986"
    ws3.column_dimensions["A"].width = 120

    ws3.merge_cells("A1:A1")
    t3 = ws3["A1"]
    t3.value = "RAW PERPLEXITY SONAR RESPONSE"
    t3.font = Font(bold=True, color="F0C040", size=13)
    t3.fill = PatternFill("solid", fgColor="0a1628")
    t3.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 28

    ws3.row_dimensions[2].height = 8

    raw = market.get("raw_text", "No data received.")
    for line_idx, line in enumerate(raw.split("\n")):
        r = line_idx + 3
        c = ws3.cell(row=r, column=1)
        c.value = line
        c.font = Font(color="c9d6e3", size=10)
        c.fill = PatternFill("solid", fgColor="0d1b2a")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[r].height = 15

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for key in ["extracted", "market", "excel_bytes", "pdf_text"]:
    if key not in st.session_state:
        st.session_state[key] = None

# ─────────────────────────────────────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("### 📄 Upload Broker Offering Memorandum")
uploaded_file = st.file_uploader(
    "Drag & drop or click to upload a PDF",
    type=["pdf"],
    label_visibility="collapsed",
)

if uploaded_file and anthropic_key:
    col_run, col_info = st.columns([1, 3])
    with col_run:
        run_btn = st.button("⚙️ Extract & Analyze", use_container_width=True)
    with col_info:
        if not perplexity_key:
            st.warning("Perplexity key not provided — market intelligence will be skipped.", icon="⚠️")

    if run_btn:
        with st.status("🔍 Processing your OM…", expanded=True) as status:

            # 1. PDF text
            st.write("📃 Extracting PDF text…")
            pdf_bytes = uploaded_file.read()
            pdf_text = extract_pdf_text(pdf_bytes)
            st.session_state["pdf_text"] = pdf_text

            # 2. Anthropic extraction
            st.write("🤖 Sending to Claude 3.5 Sonnet for structured extraction…")
            try:
                extracted = extract_with_anthropic(pdf_text, anthropic_key)
                st.session_state["extracted"] = extracted
                st.write(f"✅ Extracted {len([k for k,v in extracted.items() if v is not None])} fields.")
            except Exception as e:
                st.error(f"Anthropic extraction failed: {e}")
                st.stop()

            # 3. Perplexity
            if perplexity_key:
                city  = extracted.get("city", "")  or ""
                state = extracted.get("state", "") or ""
                ptype = extracted.get("property_type", "Commercial") or "Commercial"

                if city and state:
                    st.write(f"🌐 Fetching market intelligence for {ptype} in {city}, {state}…")
                    try:
                        market = get_perplexity_research(city, state, ptype, perplexity_key)
                        st.session_state["market"] = market
                        st.write(f"✅ Market data received. {len(market.get('comps',[]))} comps parsed.")
                    except Exception as e:
                        st.warning(f"Perplexity unavailable ({e}). Continuing without market data.")
                        st.session_state["market"] = {
                            "raw_text": f"Error: {e}", "metrics": {}, "comps": [],
                            "citations": [], "city": city, "state": state, "property_type": ptype,
                        }
                else:
                    st.warning("City/State not found in OM — skipping market research.")
                    st.session_state["market"] = None

            st.session_state["excel_bytes"] = None  # reset on new run
            status.update(label="✅ Analysis complete!", state="complete")

elif uploaded_file and not anthropic_key:
    st.warning("Please enter your Anthropic API key in the sidebar to begin.", icon="🔑")

# ─────────────────────────────────────────────────────────────────────────────
# RESULTS
# ─────────────────────────────────────────────────────────────────────────────
if st.session_state["extracted"]:
    extracted = st.session_state["extracted"]
    market    = st.session_state["market"]

    st.markdown("---")
    st.markdown("### 📊 Extraction Results")

    # ── Quick KPI row ────────────────────────────────────────────────────────
    kpis = [
        ("Property Type",    extracted.get("property_type") or "N/A",          ""),
        ("Deal Type",        extracted.get("deal_type") or "N/A",               ""),
        ("Purchase Price",   fmt_currency(extracted.get("purchase_price")),     ""),
        ("Loan Amount",      fmt_currency(extracted.get("loan_amount")),        ""),
        ("LTV",              fmt_percent(extracted.get("ltv")),                 ""),
        ("NOI",              fmt_currency(extracted.get("noi")),                extracted.get("source_noi_type") or ""),
    ]
    cols = st.columns(len(kpis))
    for col, (label, value, sub) in zip(cols, kpis):
        flag = extracted.get(f"{label.lower().replace(' ','_')}_confidence_flag","")
        badge = '<span class="flag-badge">⚠ Low</span>' if flag else ""
        col.markdown(
            f'<div class="metric-card">'
            f'<div class="label">{label}{badge}</div>'
            f'<div class="value">{value}</div>'
            f'<div class="sub">{sub}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Expanders ────────────────────────────────────────────────────────────
    with st.expander("🔍 Full Extracted PDF Data (review before export)", expanded=False):
        # highlight confidence flags
        clean = {k: v for k, v in extracted.items() if not k.endswith("_confidence_flag")}
        flags = {k: v for k, v in extracted.items() if k.endswith("_confidence_flag")}

        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Extracted Fields**")
            st.json(clean)
        with c2:
            if flags:
                st.markdown("**⚠ Confidence Flags**")
                st.json(flags)
            else:
                st.success("No low-confidence fields detected.", icon="✅")

    if market:
        with st.expander("🌐 Perplexity Market Intelligence (review before export)", expanded=False):
            m1, m2 = st.columns(2)
            with m1:
                st.markdown("**Key Market Metrics**")
                if market.get("metrics"):
                    for k, v in market["metrics"].items():
                        label = k.replace("_", " ").title()
                        st.markdown(f"- **{label}:** {v or 'N/A'}")
                else:
                    st.info("Metrics could not be parsed from response.")

                if market.get("comps"):
                    st.markdown(f"**{len(market['comps'])} Sales Comps Found**")
                    comp_df = pd.DataFrame(market["comps"])
                    st.dataframe(comp_df, use_container_width=True, hide_index=True)

            with m2:
                st.markdown("**Raw Sonar Response**")
                st.markdown(
                    f'<div style="background:#0d1b2a;border:1px solid #1e3a5f;border-radius:8px;'
                    f'padding:14px;max-height:340px;overflow-y:auto;font-size:12px;'
                    f'color:#c9d6e3;white-space:pre-wrap;">{market.get("raw_text","")}</div>',
                    unsafe_allow_html=True,
                )

    # ── Generate Excel ────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Export Institutional Memo")

    gen_col, dl_col = st.columns([1, 2])
    with gen_col:
        if st.button("📊 Generate Institutional Memo", use_container_width=True):
            if not market:
                # create empty market stub
                market = {
                    "raw_text": "Market data not available.",
                    "metrics": {},
                    "comps": [],
                    "citations": [],
                    "city": extracted.get("city", ""),
                    "state": extracted.get("state", ""),
                    "property_type": extracted.get("property_type", ""),
                }
            with st.spinner("Building Excel workbook…"):
                excel_bytes = build_excel(extracted, market)
                st.session_state["excel_bytes"] = excel_bytes
            st.success("Excel workbook ready for download!", icon="✅")

    with dl_col:
        if st.session_state["excel_bytes"]:
            prop_slug = re.sub(r"[^a-zA-Z0-9]", "_", extracted.get("property_name","OM") or "OM")[:30]
            st.download_button(
                label="⬇️  Download Institutional Memo (.xlsx)",
                data=st.session_state["excel_bytes"],
                file_name=f"RE_Debt_Memo_{prop_slug}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# ─────────────────────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<div style="text-align:center;color:#3d5570;font-size:12px;">'
    "Real Estate Debt Fund · Demo Sandbox · Powered by Claude &amp; Perplexity · Zero-Retention Architecture"
    "</div>",
    unsafe_allow_html=True,
)
